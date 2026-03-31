import random
import re
import time
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
try:
    from timetable_automation.core import load_time_slots, save_workbook_with_fallback
except ModuleNotFoundError:
    from core import load_time_slots, save_workbook_with_fallback

random.seed(42)

days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
excluded = ["07:30-09:00", "10:30-10:45", "13:15-14:00","17:30-18:30"]
# Tracks which course is using C004 in each slot (across all years/branches)
c004_occupancy = {d: {} for d in days}   # day -> {slot -> course_code}
# never allow any placement in these slots (hard ban)
ABSOLUTELY_FORBIDDEN_SLOTS = {"07:30-09:00"}


colors = [
    "FFB3BA","BAE1FF","BAFFC9","FFFFBA","FFD8BA","E3BAFF","D0BAFF","FFCBA4",
    "C7FFD8","B8E1FF","F7FFBA","FFDFBA","E9BAFF","BAFFD9","FFE1BA","BAFFF2",
    "D1FFBA","B2D8F7","F2C2FF","C2FFD8","FFB8E1","D8FFB8","FFE3BA","BAE7FF",
    "E8BAFF","BAFFD6","FFF2BA","DAD7FF","BFFFE1","FFDAB8","E2FFBA","BAF7FF"
]

thin = Border(left=Side(style='thin'), right=Side(style='thin'),
              top=Side(style='thin'), bottom=Side(style='thin'))

BASE_DIR = Path(__file__).resolve().parents[1]
DATA_DIR = BASE_DIR / "data"


def _normalize_course_dataframe(df):
    """Normalize course CSV headers across old and new semester formats."""
    rename_map = {}
    for col in df.columns:
        low = str(col).strip().lower()
        if low == "course_code":
            rename_map[col] = "Course_Code"
        elif low == "course_title":
            rename_map[col] = "Course_Title"
        elif low in {"l-t-p-s-c", "l_t_p_s_c", "ltp"}:
            rename_map[col] = "L-T-P-S-C"
        elif low == "faculty":
            rename_map[col] = "Faculty"
        elif low == "semester_half":
            rename_map[col] = "Semester_Half"
        elif low == "elective":
            rename_map[col] = "Elective"
        elif low in {"electivebasket", "elective_basket", "basket"}:
            rename_map[col] = "ElectiveBasket"
        elif low in {"is_combined", "iscombined"}:
            rename_map[col] = "Is_Combined"

    if rename_map:
        df = df.rename(columns=rename_map)

    if "Semester_Half" not in df.columns:
        df["Semester_Half"] = 0
    if "Elective" not in df.columns:
        df["Elective"] = 0
    if "ElectiveBasket" not in df.columns:
        df["ElectiveBasket"] = 0
    if "Is_Combined" not in df.columns:
        df["Is_Combined"] = 0

    return df


def _load_course_records(primary_name, *fallback_names):
    """Load the first available CSV among candidates from the data directory."""
    candidates = (primary_name,) + fallback_names
    for name in candidates:
        path = DATA_DIR / name
        if path.exists():
            df = pd.read_csv(path)
            df = _normalize_course_dataframe(df)
            return df.to_dict(orient="records"), path
    raise FileNotFoundError(
        f"None of the required data files were found: {', '.join(candidates)}"
    )


def _load_optional_course_records(*candidate_names):
    """Load optional courses; returns empty list when no candidate file exists."""
    for name in candidate_names:
        path = DATA_DIR / name
        if path.exists():
            df = pd.read_csv(path)
            df = _normalize_course_dataframe(df)
            return df.to_dict(orient="records"), path
    return [], None


def _semester_suffix(path_obj, odd_suffix, even_suffix):
    if path_obj is None:
        return odd_suffix
    return even_suffix if f"-{even_suffix}" in path_obj.name else odd_suffix


slots_norm, slot_keys, slot_dur = load_time_slots(DATA_DIR / "time_slots.json")
# elective_slots_by_year[year_tag][day] = set of slot_keys used by electives (L/T) of that year
elective_slots_by_year = {}

coursesAI, courses_ai_path = _load_course_records("coursesCSEA-I.csv", "coursesCSEA-II.csv")
coursesBI, courses_bi_path = _load_course_records("coursesCSEB-I.csv", "coursesCSEB-II.csv")
coursesA, courses_a_path = _load_course_records("coursesCSEA-III.csv", "coursesCSEA-IV.csv")
coursesB, courses_b_path = _load_course_records("coursesCSEB-III.csv", "coursesCSEB-IV.csv")

courses_v_a_path = None
courses_v_b_path = None
if (DATA_DIR / "coursesCSE-V.csv").exists():
    coursesV, courses_v_path = _load_course_records("coursesCSE-V.csv")
else:
    courses_v_a, courses_v_a_path = _load_course_records("coursesCSEA-VI.csv")
    courses_v_b, courses_v_b_path = _load_course_records("coursesCSEB-VI.csv")
    coursesV = courses_v_a + courses_v_b
    courses_v_path = courses_v_a_path

coursesDSAI, courses_dsai_path = _load_course_records("coursesDSAI-III.csv", "coursesDSAI-IV.csv")
coursesECE, courses_ece_path = _load_course_records("coursesECE-III.csv", "coursesECE-IV.csv")
coursesVII, courses_vii_path = _load_optional_course_records("courses7.csv")
coursesDSAI_I, courses_dsai_i_path = _load_course_records("coursesDSAI-I.csv", "coursesDSAI-II.csv")
coursesDSAI_V, courses_dsai_v_path = _load_course_records("coursesDSAI-V.csv", "coursesDSAI-VI.csv")
coursesECE_I, courses_ece_i_path = _load_course_records("coursesECE-I.csv", "coursesECE-II.csv")
coursesECE_V, courses_ece_v_path = _load_course_records("coursesECE-V.csv", "coursesECE-VI.csv")

SEM1_SUFFIX = _semester_suffix(courses_ai_path, "I", "II")
SEM3_SUFFIX = _semester_suffix(courses_a_path, "III", "IV")
SEM5_SUFFIX = _semester_suffix(courses_dsai_v_path or courses_v_path, "V", "VI")
YEAR1_TAG = 1 if SEM1_SUFFIX == "I" else 2
YEAR3_TAG = 3 if SEM3_SUFFIX == "III" else 4
YEAR5_TAG = 5 if SEM5_SUFFIX == "V" else 6

rooms = pd.read_csv(DATA_DIR / "rooms.csv")
rooms["Room_ID"] = rooms["Room_ID"].astype(str).str.strip()
cls = rooms[rooms["Room_ID"].str.startswith('C')].copy()
labs = rooms[rooms["Room_ID"].str.startswith('L')].copy()

try:
    reg = pd.read_csv(BASE_DIR / "registrations.csv")
    reg.set_index("Course_Code", inplace=True)
except Exception:
    reg = None

def regd(c):
    try:
        return int(reg.at[c, "Registered"])
    except Exception:
        return 0

def s(v):
    if v is None: return ""
    if isinstance(v, float) and pd.isna(v): return ""
    return str(v).strip()

def ltp(sv):
    try:
        p = [x.strip() for x in sv.split("-")]
    except Exception:
        return [0,0,0,0,0]
    while len(p) < 5:
        p.append("0")
    return list(map(int, p[:5]))

pat = re.compile(r"^[A-Z]{1,5}\d{0,3}([+/\\-][A-Z]{1,5}\d{0,3})*$", re.I)
def valid(c):
    codes, err = [], []
    for x in c:
        code = s(x.get("Course_Code", ""))
        if not code: continue
        if code.upper() in {"NEW", "ELECTIVE"}:
            codes.append(code.upper()); continue
        if not pat.match(code):
            err.append(code)
        codes.append(code.upper())
    dup = {x for x in codes if codes.count(x) > 1 and x not in {"NEW", "ELECTIVE"}}
    if dup: err += list(dup)
    return err
def is_combined_course(code, rm):
    return (code, "L") in rm and rm[(code, "L")] == "C004"
lab_prefix_for_class_prefix = {
    "C1": "L1",
    "C2": "L2",
    "C3": "L3",
    "C4": "L4",
}

def room_candidates(lab=False, prefix=None, lab_prefix=None):
    df = labs if lab else cls
    if df.empty:
        return []
    cand = df.copy()
    if prefix:
        c = cand[cand['Room_ID'].str.upper().str.startswith(prefix.upper())]
        if not c.empty:
            cand = c
        else:
            cand = df.copy()
    if lab and lab_prefix:
        c = cand[cand['Room_ID'].str.upper().str.startswith(lab_prefix.upper())]
        if not c.empty:
            cand = c
    return cand["Room_ID"].tolist()

def pick_room_for_slots(candidates, day, slots_to_use, room_busy, rr_state_key=None, rr_state=None):
    if not candidates:
        return None
    ordered = candidates
    if rr_state is not None and rr_state_key is not None and len(candidates) > 0:
        idx = rr_state.get(rr_state_key, 0) % len(candidates)
        ordered = candidates[idx:] + candidates[:idx]
    for cand in ordered:
        used = room_busy.get(day, {}).get(cand, set())
        if not (set(slots_to_use) & used):
            if rr_state is not None and rr_state_key is not None and len(candidates) > 0:
                rr_state[rr_state_key] = (rr_state.get(rr_state_key, 0) + 1) % len(candidates)
            return cand
    return None

def free(tt, d, ex=False):
    fb, b = [], []
    for s_ in slot_keys:
        if not ex and s_ in excluded:
            if b:
                fb.append(b); b = []
            continue
        if tt.at[d, s_] == "":
            b.append(s_)
        else:
            if b:
                fb.append(b); b = []
    if b: fb.append(b)
    return fb

def alloc_specific(tt, busy, rm, room_busy, day, slots_to_use, f, code, typ, elec, labsd, course_usage,
                   class_prefix=None, rr_state=None, hide_c004=False, skip_usage_check=False, ex=False, year_tag=None):
    for s_ in slots_to_use:
        if s_ not in slot_keys or tt.at[day, s_] != "":
            return False

    if code not in course_usage[day]:
        course_usage[day][code] = {"L":0,"T":0,"P":0}

    usage = course_usage[day][code]

    # For electives: Do NOT treat P as a real lab hour
    if not skip_usage_check:
        if typ == "P" and elec:
            # elective lab behaves like theory — allow unlimited placement
            pass
        else:
            if typ == "P":
                if usage["P"] >= 1:
                    return False
            else:
                if (usage["L"] + usage["T"]) >= 1:
                    return False


    r = None
    if not elec:
        key = (code, typ)
        if key in rm:
            candidate = rm[key]
            # if candidate is C004 we still need to check cross-branch occupancy below
            if candidate != "C004":
                used = room_busy.get(day, {}).get(candidate, set())
                if set(slots_to_use) & used:
                    return False
            r = candidate
        else:
            if typ == "P" and elec:
                r = None
            elif typ == "P":
                lab_pref = lab_prefix_for_class_prefix.get(class_prefix, None)
                candidates = room_candidates(lab=True, prefix=None, lab_prefix=lab_pref)
            else:
                candidates = room_candidates(lab=False, prefix=class_prefix, lab_prefix=None)
            r = pick_room_for_slots(candidates, day, slots_to_use, room_busy, rr_state_key=class_prefix, rr_state=rr_state)
            if r is None:
                return False
            rm[key] = r

    # NEW: Block C004 if occupied by a different course for any requested slot
    if r == "C004":
        for s_ in slots_to_use:
            occ = c004_occupancy.get(day, {}).get(s_)
            if occ and occ != code:
                # slot in C004 already taken by a different course
                return False

    # Commit the allocation to tt
    for s_ in slots_to_use:
        if is_combined_course(code, rm):
            if hide_c004:
                if typ == "P":
                    v = f"{code} (Lab)"
                elif typ == "T":
                    v = f"{code}T"
                else:
                    v = f"{code}"
            else:
                if typ == "P":
                    v = f"{code} (Lab)"
                elif typ == "T":
                    v = f"{code}T (C004)"
                else:
                    v = f"{code} (C004)"
        else:
            if r and not elec:
                if elec and typ == "P":
                    v = f"{code}(Lab)"
                elif typ == "T":
                    v = f"{code}T ({r})"
                elif typ == "P":
                    v = f"{code} (Lab-{r})"
                else:
                    v = f"{code} ({r})"
            else:
                if elec and typ == "P":
                    v = f"{code}(Lab)"
                elif typ == "T":
                    v = f"{code}T"
                else:
                    v = code
        tt.at[day, s_] = v

    if f:
        busy[day].setdefault(f, set()).update(slots_to_use)
    if r:
        room_busy.setdefault(day, {}).setdefault(r, set()).update(slots_to_use)
    if typ == "P":
        labsd.add(day)
    course_usage[day][code][typ] += 1

    # NEW: mark C004 occupancy so other branches see it
    if r == "C004":
        for s_ in slots_to_use:
            c004_occupancy.setdefault(day, {})[s_] = code

    return True



def alloc(tt, busy, rm, room_busy, d, f, code, h, typ="L", elec=False, labsd=set(), ex=False,
          preferred_slots=None, course_usage=None, class_prefix=None, rr_state=None, hide_c004=False,year_tag=None):
    if course_usage is None:
        course_usage = {dd:{} for dd in days}
    if code not in course_usage[d]:
        course_usage[d][code] = {"L":0,"T":0,"P":0}

    usage = course_usage[d][code]

    if typ == "P":
        if usage["P"] >= 1:
            return False
    else:
        if (usage["L"] + usage["T"]) >= 1:
            return False

    if preferred_slots:
        pref_day, pref_slots = preferred_slots
        if pref_day == d:
            total = sum(slot_dur[s] for s in pref_slots)
            if total + 1e-9 >= h:
                if alloc_specific(tt, busy, rm, room_busy, pref_day, pref_slots, f, code, typ, elec, labsd, course_usage, class_prefix=class_prefix, rr_state=rr_state, hide_c004=hide_c004):
                    return True

    for blk in free(tt, d, ex):
        if sum(slot_dur[s] for s in blk) + 1e-9 < h: continue
        use = []; dur = 0.0
        for s_ in blk:
            use.append(s_); dur += slot_dur[s_]
            if dur + 1e-9 >= h: break
        if not ex and any(s_ in excluded for s_ in use): continue
        if f and f in busy[d] and (set(use) & busy[d][f]): continue

        if not elec:
            key = (code, typ)
            if key in rm:
                r = rm[key]
                if r != "C004":
                    used = room_busy.get(d, {}).get(r, set())
                    if set(use) & used:
                        continue
            else:
                if typ == "P" and elec:
                    r = None
                elif typ == "P":
                    lab_pref = lab_prefix_for_class_prefix.get(class_prefix, None)
                    candidates = room_candidates(lab=True, prefix=None, lab_prefix=lab_pref)
                    r = pick_room_for_slots(candidates, d, use, room_busy, rr_state_key=lab_pref, rr_state=rr_state)
                else:
                    candidates = room_candidates(lab=False, prefix=class_prefix, lab_prefix=None)
                    r = pick_room_for_slots(candidates, d, use, room_busy, rr_state_key=class_prefix, rr_state=rr_state)

                if r is None:
                    continue
                rm[(code, typ)] = r
        else:
            r = None

        # NEW: prevent C004 being used by a different course simultaneously
        if r == "C004":
            conflict = False
            for s_ in use:
                occ = c004_occupancy.get(d, {}).get(s_)
                if occ and occ != code:
                    conflict = True; break
            if conflict:
                continue

        # commit allocation to cells
        for s_ in use:
            if is_combined_course(code, rm):
                if hide_c004:
                    if typ == "P":
                        v = f"{code}(Lab)"
                    elif typ == "T":
                        v = f"{code}T"
                    else:
                        v = f"{code}"
                else:
                    if typ == "P":
                        v = f"{code} (Lab)"
                    elif typ == "T":
                        v = f"{code}T (C004)"
                    else:
                        v = f"{code} (C004)"
            else:
                if r and not elec:
                    if elec and typ == "P":
                        v = f"{code}(Lab)"
                    elif typ == "T":
                        v = f"{code}T ({r})"
                    elif typ == "P":
                        v = f"{code} (Lab-{r})"
                    else:
                        v = f"{code} ({r})"
                else:
                    if elec and typ == "P":
                        v = f"{code}(Lab)"
                    elif typ == "T":
                        v = f"{code}T"
                    else:
                        v = code
            tt.at[d, s_] = v

        if f:
            busy[d].setdefault(f, set()).update(use)
        if r:
            room_busy.setdefault(d, {}).setdefault(r, set()).update(use)
        if typ == "P":
            labsd.add(d)
        course_usage[d][code][typ] += 1

        # NEW: mark C004 occupancy
        if r == "C004":
            for s_ in use:
                c004_occupancy.setdefault(d, {})[s_] = code

        return True

    return False


def get_all_valid_free_slots(tt):
    valid = []
    for d in reversed(days):
        for s_ in reversed(slot_keys):
            if s_ in excluded: continue
            if tt.at[d, s_] == "": valid.append((d, s_))
    return valid

def get_all_excluded_free_slots(tt):
    exs = []
    for d in reversed(days):
        for s_ in reversed(slot_keys):
            if s_ not in excluded: continue
            if tt.at[d, s_] == "": exs.append((d, s_))
    return exs

def extract_contiguous_blocks(slot_list):
    blocks = []
    i = 0
    while i < len(slot_list):
        d0, s0 = slot_list[i]
        cur_day = d0
        cur_slots = [s0]
        i += 1
        while i < len(slot_list) and slot_list[i][0] == cur_day:
            cur_slots.append(slot_list[i][1]); i += 1
        blocks.append((cur_day, cur_slots))
    return blocks

def try_allocate_chunk_from_block(
    tt, busy, rm, room_busy, labsd, course_usage,
    code, faculty, typ, need, day, slots,
    class_prefix=None, rr_state=None, hide_c004=False,
    allow_excluded=False
):
    n = len(slots)

    best_sub = None
    best_i = best_j = None
    best_total = None

    # 1️⃣ Find best contiguous slice
    for i in range(n):
        accum = 0.0
        sub = []          # ✅ MUST be here

        for j in range(i, n):
            s_key = slots[j]

            # 🚫 Hard-forbid excluded slots
            if s_key in excluded and not allow_excluded:
                break

            sub.append(s_key)
            accum += slot_dur[s_key]

            if accum + 1e-9 >= need:
                if all(tt.at[day, s_] == "" for s_ in sub):
                    if best_sub is None or accum < best_total - 1e-9:
                        best_sub = list(sub)
                        best_i, best_j, best_total = i, j, accum
                break

    if best_sub is None:
        return None, None

    # 2️⃣ Allocate chosen slice
    ok = alloc_specific(
        tt, busy, rm, room_busy,
        day, best_sub,
        faculty, code, typ,
        False, labsd, course_usage,
        class_prefix=class_prefix,
        rr_state=rr_state,
        hide_c004=hide_c004
    )

    if not ok:
        return None, None

    new_slots = slots[:best_i] + slots[best_j + 1:]
    return new_slots, best_sub




def get_free_blocks_for_combined(tt):
    """
    Recompute all free (non-excluded) contiguous blocks for combined courses.
    Returns list of (day, [slot_keys...]) where slots are free and contiguous.
    """
    blocks = []
    for d in days:
        cur = []
        for s_ in slot_keys:
            if s_ in excluded:
                if cur:
                    blocks.append((d, cur))
                    cur = []
                continue
            if tt.at[d, s_] == "":
                cur.append(s_)
            else:
                if cur:
                    blocks.append((d, cur))
                    cur = []
        if cur:
            blocks.append((d, cur))
    return blocks

def find_contiguous_slice(block_slots, need_hours):
    """
    Given a list of free slots on a single day, find a contiguous
    sub-slice whose total duration >= need_hours. Returns the slice
    (list of slot_keys) or None.
    """
    n = len(block_slots)
    for i in range(n):
        accum = 0.0
        sub = []
        for j in range(i, n):
            s_ = block_slots[j]
            accum += slot_dur[s_]
            sub.append(s_)
            if accum + 1e-9 >= need_hours:
                return sub
    return None


def assign_combined_precise_durations(
    tt, busy, rm, room_busy, labsd, course_usage, combined_core,
    rr_state=None, hide_c004=False,
    combined_sync=None, year_tag=None, semester_half=None
):
    ALLOWED_LECTURE_CHUNKS = [1.5, 1.0]

    if not combined_core:
        return []

    combined_list = []
    chunks_map = {}

    # ---------- BUILD CHUNKS ----------
    for c in combined_core:
        code = s(c.get("Course_Code", ""))
        if not code:
            continue

        rm[(code, "L")] = "C004"
        rm[(code, "T")] = "C004"
        rm[(code, "P")] = "C004"

        L, T, P, _, _ = ltp(c.get("L-T-P-S-C", "0-0-0-0-0"))

        ch = []

        rem = float(L)
        while rem > 1e-9:
            if rem >= 1.5:
                ch.append((1.5, "L"))
                rem -= 1.5
            else:
                ch.append((1.0, "L"))
                rem -= 1.0

        rem = float(T)
        while rem > 1e-9:
            ch.append((1.0, "T"))
            rem -= 1.0

        rem = float(P)
        while rem > 1e-9:
            if rem >= 2.0:
                ch.append((2.0, "P"))
                rem -= 2.0
            elif rem >= 1.5:
                ch.append((1.5, "P"))
                rem -= 1.5
            else:
                ch.append((1.0, "P"))
                rem -= 1.0

        chunks_map[code] = sorted(ch, key=lambda x: -x[0])
        combined_list.append((code, c))

    combined_list.sort(key=lambda x: x[0])
    placed_codes = []

    # ---------- PLACEMENT ----------
    for code, c in combined_list:
        faculty = s(c.get("Faculty", ""))
        chunks = chunks_map[code]

        sync_key = None
        existing_sync = None
        if combined_sync is not None and year_tag is not None:
            sync_key = (year_tag, code)
            existing_sync = combined_sync.get(sync_key)

        new_sync_entries = []
        days_used = set()

        for idx, (need, typ) in enumerate(chunks):
            allocated = False

            # ===== LECTURE-SPECIFIC LOGIC =====
            if typ == "L":
                remaining = need

                while remaining > 1e-6:
                    placed = False

                    valid_slots = get_all_valid_free_slots(tt)
                    valid_blocks = extract_contiguous_blocks(valid_slots)

                    for chunk in ALLOWED_LECTURE_CHUNKS:
                        if chunk > remaining + 1e-6:
                            continue

                        for day, slots in valid_blocks:
                            if day in days_used:
                                continue

                            _, used_slots = try_allocate_chunk_from_block(
                                tt, busy, rm, room_busy, labsd, course_usage,
                                code, faculty, "L", chunk, day, slots,
                                class_prefix="C0",
                                rr_state=rr_state,
                                hide_c004=hide_c004
                            )

                            if used_slots is not None:
                                new_sync_entries.append(("L", day, used_slots))
                                days_used.add(day)
                                remaining -= chunk
                                placed = True
                                break

                        if placed:
                            break

                    if not placed:
                        break

                allocated = True
                continue   # ⛔ do NOT fall into mirror / generic logic

            # ===== MIRROR LOGIC =====
            if existing_sync is not None and idx < len(existing_sync):
                sync_typ, sync_day, sync_slots = existing_sync[idx]
                if sync_typ == typ:
                    ok = alloc_specific(
                        tt, busy, rm, room_busy,
                        sync_day, sync_slots,
                        None,
                        code, typ, False, labsd, course_usage,
                        class_prefix="C0",
                        rr_state=rr_state,
                        hide_c004=hide_c004
                    )
                    if ok:
                        allocated = True
                        days_used.add(sync_day)

            # ===== NORMAL BLOCK SEARCH =====
            if not allocated:
                valid_slots = get_all_valid_free_slots(tt)
                valid_blocks = extract_contiguous_blocks(valid_slots)

                for day, slots in valid_blocks:
                    if day in days_used:
                        continue

                    _, used_slots = try_allocate_chunk_from_block(
                        tt, busy, rm, room_busy, labsd, course_usage,
                        code, faculty, typ, need, day, slots,
                        class_prefix="C0",
                        rr_state=rr_state,
                        hide_c004=hide_c004
                    )

                    if used_slots is not None:
                        new_sync_entries.append((typ, day, used_slots))
                        allocated = True
                        days_used.add(day)
                        break

            if not allocated:
                break

        if sync_key and combined_sync is not None and existing_sync is None and new_sync_entries:
            combined_sync[sync_key] = new_sync_entries

        placed_codes.append(code)

    return placed_codes



color_avail = colors.copy(); random.shuffle(color_avail); color_map = {}
def get_color_for_course(course_code):
    k = course_code.strip().upper()
    if k == "": return None
    if k not in color_map:
        if color_avail: color_map[k] = color_avail.pop()
        else: color_map[k] = "CCCCCC"
    return color_map[k]

def merge_and_color(ws, courses):
    sc = 2
    mc = ws.max_column
    mr = ws.max_row

    valid_course_codes = {
        s(x.get("Course_Code", "")).replace("T", "").strip().upper()
        for x in courses
        if s(x.get("Course_Code", ""))
    }
    valid_course_codes |= {f"ELECTIVE{i}" for i in range(1, 60)}

    # Header styling
    for col in range(2, mc + 1):
        cell = ws.cell(2, col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin

    # Merge contiguous identical entries in each row
    for r in range(3, mr + 1):
        c = sc
        while c <= mc:
            raw = ws.cell(r, c).value
            if raw is None or str(raw).strip() == "":
                ws.cell(r, c).border = thin
                c += 1
                continue

            val = str(raw).strip()
            merge_cols = [c]

            # extend to all immediately-adjacent cells with same text
            next_col = c + 1
            while next_col <= mc:
                next_raw = ws.cell(r, next_col).value
                next_val = str(next_raw).strip() if next_raw is not None else ""
                if next_val == val:
                    merge_cols.append(next_col)
                    next_col += 1
                else:
                    break

            # actually merge the block if it spans >1 column
            if len(merge_cols) > 1:
                ws.merge_cells(
                    start_row=r,
                    start_column=merge_cols[0],
                    end_row=r,
                    end_column=merge_cols[-1]
                )

            # styling + colour
            cell = ws.cell(r, merge_cols[0])
            cell.alignment = Alignment(horizontal="center",
                                       vertical="center",
                                       wrap_text=True)
            cell.font = Font(bold=True)

            raw_course_name = val.split()[0] if val.split() else val
            raw_course_name = (
                raw_course_name.replace("T", "")
                .replace("(", "")
                .strip()
                .upper()
            )
            fill_color = (
                get_color_for_course(raw_course_name)
                if (raw_course_name in valid_course_codes
                    or raw_course_name.startswith("ELECTIVE"))
                else None
            )

            for cc_ in merge_cols:
                cell_ref = ws.cell(r, cc_)
                cell_ref.border = thin
                cell_ref.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )
                cell_ref.font = Font(bold=True)
                if fill_color:
                    cell_ref.fill = PatternFill(
                        start_color=fill_color,
                        end_color=fill_color,
                        fill_type="solid"
                    )

            c = merge_cols[-1] + 1

    # auto column widths
    for col in ws.columns:
        maxl = 0
        cl = col[0].column_letter
        for cell in col:
            v = cell.value
            if v is None:
                continue
            maxl = max(maxl, len(str(v)))
        ws.column_dimensions[cl].width = min(maxl + 2 if maxl else 8, 60)

def add_csv_legend_block(ws, csv_path, legend_title, room_prefix=None, elective_room_map=None):
    if elective_room_map is None:
        elective_room_map = {}

    ws.append([""]); ws.append([""]); ws.append([f"Legend - {legend_title}"])
    title_cell = ws.cell(row=ws.max_row, column=1)
    title_cell.font = Font(bold=True, size=13)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    df = _normalize_course_dataframe(pd.read_csv(csv_path))
    expect_cols = ["Course_Code", "Course_Title", "L-T-P-S-C", "Faculty", "Semester_Half", "Elective", "ElectiveBasket"]
    for ec in expect_cols:
        if ec not in df.columns:
            alt = None; low = ec.lower()
            for c in df.columns:
                if c.lower() == low:
                    alt = c; break
            if alt: df.rename(columns={alt: ec}, inplace=True)
            else:
                if ec == "Semester_Half": df[ec] = 0
                elif ec == "Elective": df[ec] = 0
                else: df[ec] = ""

    df = df[["Course_Code", "Course_Title", "L-T-P-S-C", "Faculty", "Semester_Half", "Elective", "ElectiveBasket"]].copy()

    def map_sem(x):
        try: xi = int(x)
        except Exception: xi = 0
        if xi == 1: return "First Half"
        if xi == 2: return "Second Half"
        return "Full Sem"
    def map_elec(x):
        try: xi = int(x)
        except Exception: xi = 0
        return "Yes" if xi == 1 else "No"

    df["Semester_Half"] = df["Semester_Half"].apply(map_sem)
    df["Elective"] = df["Elective"].apply(map_elec)

    all_classrooms = cls["Room_ID"].tolist()

    master_pool = sorted(list(set(all_classrooms)))
    random.shuffle(master_pool)
    # ------------------------------------------------------

    elective_rooms = []
    for _, row in df.iterrows():
        if row["Elective"] == "Yes":
            basket = str(row.get("ElectiveBasket", "")).strip()
            if basket and basket != "0":
                sync_name = f"{row['Course_Code']}_B{basket}"
            else:
                sync_name = row["Course_Code"]

            if sync_name in elective_room_map:
                chosen = elective_room_map[sync_name]
            else:
                taken_rooms = set(elective_room_map.values())
            
                candidates = [r for r in master_pool if r not in taken_rooms]
                
                if candidates:
                    chosen = candidates[0]
                else:
                    chosen = random.choice(master_pool)
                
                elective_room_map[sync_name] = chosen

            elective_rooms.append(f"{chosen}")
        else:
            elective_rooms.append("")

    df["Elective Room"] = elective_rooms

    headers = ["Course Code","Course Title","L-T-P-S-C","Faculty","Semester Half","Elective","Elective Basket","Elective Room"]

    ws.append(headers); header_row = ws.max_row
    for i, _h in enumerate(headers, start=1):
        c = ws.cell(header_row, i); c.font = Font(bold=True); c.alignment = Alignment(horizontal="center", vertical="center"); c.border = thin; c.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    for idx, row in df.iterrows():
        rowvals = [
            s(row["Course_Code"]),
            s(row["Course_Title"]),
            s(row["L-T-P-S-C"]),
            s(row["Faculty"]),
            s(row["Semester_Half"]),
            s(row["Elective"]),
            s(row["ElectiveBasket"]),
            row["Elective Room"]
        ]
        ws.append(rowvals)
        for i in range(1, 8):
            cc = ws.cell(ws.max_row, i); cc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); cc.border = thin
    ws.append([""])

def generate(courses, ws, label, seed, elective_sync,
             room_prefix=None, elective_room_map=None,
             room_busy_global=None, hide_c004=False,
             year_tag=None, combined_sync=None,semester_half=None):
    if elective_room_map is None:
        elective_room_map = {}
    if valid(courses): return []
    
    ws.append([""]); ws.append([label])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
    
    tt = pd.DataFrame("", index=days, columns=slot_keys)
    busy = {d:{} for d in days}
    
    if room_busy_global is not None:
        room_busy = room_busy_global
    else:
        room_busy = {d:{} for d in days}

    rm = {}
    labsd = set()
    course_usage = {d:{} for d in days}
    rr_state = {}

    elec = [x for x in courses if s(x.get("Elective","")) == "1"]
    combined_core = [x for x in courses if s(x.get("Elective","")) != "1" and s(x.get("Is_Combined","0")) == "1"]
    regular_core = [x for x in courses if s(x.get("Elective","")) != "1" and s(x.get("Is_Combined","0")) != "1"]

    baskets = {}; elec_no_baskets = []
    for e in elec:
        b = s(e.get("ElectiveBasket","0"))
        if b and b != "0": baskets.setdefault(b,[]).append(e)
        else: elec_no_baskets.append(e)
    basket_reps = []
    for b, group in sorted(baskets.items(), key=lambda x: int(x[0]) if x[0].isdigit() else 0):
        chosen = group[0]
        sync_identifier = f"{chosen.get('Course_Code')}_B{b}"
        basket_reps.append({
            "Course_Code": f"Elective{b}",
            "Course_Title": chosen.get("Course_Title","") or chosen.get("Course_Code",""),
            "Faculty": chosen.get("Faculty",""),
            "L-T-P-S-C": chosen.get("L-T-P-S-C","0-0-0-0-0"),
            "Elective": "1",
            "ElectiveBasket": b,
            "_sync_name": sync_identifier
        })

    for e in elec_no_baskets:
        basket = s(e.get("ElectiveBasket","0"))
        if basket and basket != "0":
            sync_n = f"{s(e.get('Course_Code'))}_B{basket}"
        else:
            sync_n = s(e.get("Course_Code"))
        e["_sync_name"] = sync_n if sync_n else None
    elec_final = elec_no_baskets + basket_reps

    for c in combined_core:
        code = s(c.get("Course_Code",""))
        rm[(code,"L")] = "C004"; rm[(code,"T")] = "C004"; rm[(code,"P")] = "C004"

    def place_course_list(course_list, start_idx_ref):
        placed_list = []
        for c in course_list:
            f = s(c.get("Faculty",""))
            code = s(c.get("Course_Code","UNKNOWN"))
            is_elec_flag = (code.startswith("Elective") or s(c.get("Elective","")) == "1")
            L, T, P, S, Cc = ltp(c.get("L-T-P-S-C","0-0-0-0-0"))
            for h, typ in [(L,"L"), (T,"T"), (P,"P")]:
                attempts = 0
                while h > 1e-9 and attempts < 400:
                    # Enforce strict durations per type:
                    # - Lecture (L) => always 1.5 hours
                    # - Tutorial (T) => always 1.0 hour
                    # - Practical/Lab (P) => prefer 2.0, else 1.5, else 1.0 (only if remaining h is smaller)
                    if typ == "L":
                        a = 1.5
                    elif typ == "T":
                        a = 1.0
                    elif typ == "P":
                        # For labs prefer 2.0 blocks; if remaining hours < 2, allow smaller lab chunk
                        if h >= 2.0 - 1e-9:
                            a = 2.0
                        elif h >= 1.5 - 1e-9:
                            a = 1.5
                        else:
                            a = 1.0
                    else:
                        a = 1.0
                    placed = False
                    sync_name = c.get("_sync_name", None)

                    if is_elec_flag and sync_name and sync_name in elective_room_map:
                        for ttkey in [("L"), ("T"), ("P")]:
                            rm[(code, ttkey)] = elective_room_map[sync_name]

                    if sync_name and sync_name in elective_sync:
                        pref = elective_sync[sync_name]
                        if alloc(tt, busy, rm, room_busy, pref["day"], f, code, a, typ, is_elec_flag, labsd, False, preferred_slots=(pref["day"], pref["slots"]), course_usage=course_usage, class_prefix=room_prefix, rr_state=rr_state,hide_c004=hide_c004,year_tag=year_tag):
                            h -= a; placed = True

                    if not placed:
                        for i in range(5):
                            if is_elec_flag:
                                d_order = days[:]
                            else:
                                start_idx = start_idx_ref[0]
                                d_order = days[start_idx:] + days[:start_idx]
                                start_idx_ref[0] = (start_idx_ref[0] + 1) % len(days)
                            for d in d_order:
                                if alloc(tt, busy, rm, room_busy, d, f, code, a, typ, is_elec_flag, labsd, False, course_usage=course_usage, class_prefix=room_prefix, rr_state=rr_state,hide_c004=hide_c004,year_tag=year_tag):
                                    h -= a; placed = True; break
                            if placed:
                                break
                    if not placed:
                        for d in days:
                            if alloc(tt, busy, rm, room_busy, d, f, code, a, typ, is_elec_flag, labsd, True, course_usage=course_usage, class_prefix=room_prefix, rr_state=rr_state,hide_c004=hide_c004,year_tag=year_tag):
                                h -= a; placed = True; break

                    if placed and sync_name and sync_name not in elective_sync:
                        for dcheck in days:
                            slots_used = [s_ for s_ in slot_keys if tt.at[dcheck, s_].startswith(code)]
                            if slots_used:
                                accum = []; acc_dur = 0.0
                                for s_ in slots_used:
                                    accum.append(s_); acc_dur += slot_dur[s_]
                                    if acc_dur + 1e-9 >= a:
                                        elective_sync[sync_name] = {"day": dcheck, "slots": accum.copy()}
                                        break
                                if sync_name in elective_sync: break

                    attempts += 1
            placed_list.append(c)
        return placed_list

    start_idx_ref = [seed % len(days)]
    elec_final.sort(key=lambda x: 0 if x.get("_sync_name") in elective_sync else 1)
    
    priority_placed = place_course_list(elec_final, start_idx_ref)

    combined_placed = assign_combined_precise_durations(
        tt, busy, rm, room_busy, labsd, course_usage, combined_core,
        rr_state=rr_state, hide_c004=hide_c004,  combined_sync=combined_sync, year_tag=year_tag,semester_half=semester_half
    )
    regular_placed = place_course_list(regular_core, start_idx_ref)

    ws.append(["Day"] + slot_keys)
    for d in days:
        ws.append([d] + [tt.at[d, s] for s in slot_keys])
    ws.append([""])
    return (priority_placed + regular_placed + combined_core)
def split(c):
    f = [x for x in c if s(x.get("Semester_Half","")) in ["1","0"]]
    s2 = [x for x in c if s(x.get("Semester_Half","")) in ["2","0"]]
    return f, s2

if __name__ == "__main__":
    wb = Workbook()
    seed = random.randint(0, 999999)

    elective_room_map = {}
    global_room_busy = {d: {} for d in days}

    sync_sem1 = {}
    sync_sem3 = {}
    sync_sem5 = {}
    sync_sem7 = {}
    combined_sync_sem1 = {}
    combined_sync_sem3 = {}

    ws1 = wb.active
    ws1.title = f"CSE-{SEM1_SUFFIX} Timetable"
    cAf, cAs = split(coursesAI)
    cBf, cBs = split(coursesBI)
    
    csea_block = generate(cAf, ws1, f"CSEA {SEM1_SUFFIX} First Half", seed+0, sync_sem1, room_prefix='C1', elective_room_map=elective_room_map, room_busy_global=global_room_busy,hide_c004=True,year_tag=YEAR1_TAG,combined_sync=combined_sync_sem1,semester_half=1)
    csea_block2 = generate(cAs, ws1, f"CSEA {SEM1_SUFFIX} Second Half", seed+1, sync_sem1, room_prefix='C1', elective_room_map=elective_room_map, room_busy_global=global_room_busy,hide_c004=True,year_tag=YEAR1_TAG,combined_sync=combined_sync_sem1,semester_half=2)
    add_csv_legend_block(ws1, str(courses_ai_path), f"CSEA {SEM1_SUFFIX}", room_prefix="C1", elective_room_map=elective_room_map)
    
    cseb_block = generate(cBf, ws1, f"CSEB {SEM1_SUFFIX} First Half", seed+2, sync_sem1, room_prefix='C1', elective_room_map=elective_room_map, room_busy_global=global_room_busy,hide_c004=True,year_tag=YEAR1_TAG,combined_sync=combined_sync_sem1,semester_half=1)
    cseb_block2 = generate(cBs, ws1, f"CSEB {SEM1_SUFFIX} Second Half", seed+3, sync_sem1, room_prefix='C1', elective_room_map=elective_room_map, room_busy_global=global_room_busy,hide_c004=True,year_tag=YEAR1_TAG,combined_sync=combined_sync_sem1,semester_half=2)
    add_csv_legend_block(ws1, str(courses_bi_path), f"CSEB {SEM1_SUFFIX}", room_prefix="C1", elective_room_map=elective_room_map)
    
    combined_i_courses = (csea_block or []) + (csea_block2 or []) + (cseb_block or []) + (cseb_block2 or [])
    merge_and_color(ws1, combined_i_courses)

    # --- DSAI (lower semester) ---
    ws7 = wb.create_sheet(f"DSAI-{SEM1_SUFFIX} Timetable")
    d1f_i, d1s_i = split(coursesDSAI_I)
    dsai1_block1 = generate(d1f_i, ws7, f"DSAI-{SEM1_SUFFIX} First Half", seed+16, sync_sem1, room_prefix='C1', elective_room_map=elective_room_map, room_busy_global=global_room_busy,year_tag=YEAR1_TAG,combined_sync=combined_sync_sem1,semester_half=1)
    dsai1_block2 = generate(d1s_i, ws7, f"DSAI-{SEM1_SUFFIX} Second Half", seed+17, sync_sem1, room_prefix='C1', elective_room_map=elective_room_map, room_busy_global=global_room_busy,year_tag=YEAR1_TAG,combined_sync=combined_sync_sem1,semester_half=2)
    add_csv_legend_block(ws7, str(courses_dsai_i_path), f"DSAI {SEM1_SUFFIX}", room_prefix="C1", elective_room_map=elective_room_map)
    combined_dsai1_courses = (dsai1_block1 or []) + (dsai1_block2 or [])
    merge_and_color(ws7, combined_dsai1_courses)

    # --- ECE (lower semester) ---
    ws9 = wb.create_sheet(f"ECE-{SEM1_SUFFIX} Timetable")
    e1f_i, e1s_i = split(coursesECE_I)
    ece1_block1 = generate(e1f_i, ws9, f"ECE-{SEM1_SUFFIX} First Half", seed+20, sync_sem1, room_prefix='C4', elective_room_map=elective_room_map, room_busy_global=global_room_busy,year_tag=YEAR1_TAG,combined_sync=combined_sync_sem1,semester_half=1)
    ece1_block2 = generate(e1s_i, ws9, f"ECE-{SEM1_SUFFIX} Second Half", seed+21, sync_sem1, room_prefix='C4', elective_room_map=elective_room_map, room_busy_global=global_room_busy,year_tag=YEAR1_TAG,combined_sync=combined_sync_sem1,semester_half=2)
    add_csv_legend_block(ws9, str(courses_ece_i_path), f"ECE {SEM1_SUFFIX}", room_prefix="C4", elective_room_map=elective_room_map)
    combined_ece1_courses = (ece1_block1 or []) + (ece1_block2 or [])
    merge_and_color(ws9, combined_ece1_courses)
    # --- CSE mid semester (Sections A & B) ---
    ws2 = wb.create_sheet(f"CSE-{SEM3_SUFFIX} Timetable")
    c1f, c1s = split(coursesA); c2f, c2s = split(coursesB)
    
    csea3_block1 = generate(c1f, ws2, f"CSEA {SEM3_SUFFIX} First Half", seed+4, sync_sem3, room_prefix='C2', elective_room_map=elective_room_map, room_busy_global=global_room_busy,year_tag=YEAR3_TAG,combined_sync=combined_sync_sem3,semester_half=1)
    csea3_block2 = generate(c1s, ws2, f"CSEA {SEM3_SUFFIX} Second Half", seed+5, sync_sem3, room_prefix='C2', elective_room_map=elective_room_map, room_busy_global=global_room_busy,year_tag=YEAR3_TAG,combined_sync=combined_sync_sem3,semester_half=2)
    add_csv_legend_block(ws2, str(courses_a_path), f"CSEA {SEM3_SUFFIX}", room_prefix="C2", elective_room_map=elective_room_map)
    
    cseb3_block1 = generate(c2f, ws2, f"CSEB {SEM3_SUFFIX} First Half", seed+6, sync_sem3, room_prefix='C2', elective_room_map=elective_room_map, room_busy_global=global_room_busy,year_tag=YEAR3_TAG,combined_sync=combined_sync_sem3,semester_half=1)
    cseb3_block2 = generate(c2s, ws2, f"CSEB {SEM3_SUFFIX} Second Half", seed+7, sync_sem3, room_prefix='C2', elective_room_map=elective_room_map, room_busy_global=global_room_busy,year_tag=YEAR3_TAG,combined_sync=combined_sync_sem3,semester_half=2)
    add_csv_legend_block(ws2, str(courses_b_path), f"CSEB {SEM3_SUFFIX}", room_prefix="C2", elective_room_map=elective_room_map)
    
    combined_iii_courses = (csea3_block1 or []) + (csea3_block2 or []) + (cseb3_block1 or []) + (cseb3_block2 or [])
    merge_and_color(ws2, combined_iii_courses)

    # --- DSAI mid semester ---
    ws4 = wb.create_sheet(f"DSAI-{SEM3_SUFFIX} Timetable")
    d1f, d1s = split(coursesDSAI)
    dsa_block1 = generate(d1f, ws4, f"DSAI-{SEM3_SUFFIX} First Half", seed+10, sync_sem3, room_prefix='C4', elective_room_map=elective_room_map, room_busy_global=global_room_busy,year_tag=YEAR3_TAG,combined_sync=combined_sync_sem3,semester_half=1)
    dsa_block2 = generate(d1s, ws4, f"DSAI-{SEM3_SUFFIX} Second Half", seed+11, sync_sem3, room_prefix='C4', elective_room_map=elective_room_map, room_busy_global=global_room_busy,year_tag=YEAR3_TAG,combined_sync=combined_sync_sem3,semester_half=2)
    add_csv_legend_block(ws4, str(courses_dsai_path), f"DSAI {SEM3_SUFFIX}", room_prefix="C4", elective_room_map=elective_room_map)
    combined_dsa_courses = (dsa_block1 or []) + (dsa_block2 or [])
    merge_and_color(ws4, combined_dsa_courses)

    # --- ECE mid semester ---
    ws5 = wb.create_sheet(f"ECE-{SEM3_SUFFIX} Timetable")
    e1f, e1s = split(coursesECE)
    ece_block1 = generate(e1f, ws5, f"ECE-{SEM3_SUFFIX} First Half", seed+12, sync_sem3, room_prefix='C4', elective_room_map=elective_room_map, room_busy_global=global_room_busy,year_tag=YEAR3_TAG,combined_sync=combined_sync_sem3,semester_half=1)
    ece_block2 = generate(e1s, ws5, f"ECE-{SEM3_SUFFIX} Second Half", seed+13, sync_sem3, room_prefix='C4', elective_room_map=elective_room_map, room_busy_global=global_room_busy,year_tag=YEAR3_TAG,combined_sync=combined_sync_sem3,semester_half=2)
    add_csv_legend_block(ws5, str(courses_ece_path), f"ECE {SEM3_SUFFIX}", room_prefix="C4", elective_room_map=elective_room_map)
    combined_ece_courses = (ece_block1 or []) + (ece_block2 or [])
    merge_and_color(ws5, combined_ece_courses)

    # --- CSE upper semester ---
    ws3 = wb.create_sheet(f"CSE-{SEM5_SUFFIX} Timetable")
    c5f, c5s = split(coursesV)
    c5_block1 = generate(c5f, ws3, f"CSE-{SEM5_SUFFIX} First Half", seed+8, sync_sem5, room_prefix='C3', elective_room_map=elective_room_map, room_busy_global=global_room_busy,year_tag=YEAR5_TAG,semester_half=1)
    c5_block2 = generate(c5s, ws3, f"CSE-{SEM5_SUFFIX} Second Half", seed+9, sync_sem5, room_prefix='C3', elective_room_map=elective_room_map, room_busy_global=global_room_busy,year_tag=YEAR5_TAG,semester_half=2)
    if courses_v_a_path is not None and courses_v_b_path is not None:
        add_csv_legend_block(ws3, str(courses_v_a_path), f"CSEA {SEM5_SUFFIX}", room_prefix="C3", elective_room_map=elective_room_map)
        add_csv_legend_block(ws3, str(courses_v_b_path), f"CSEB {SEM5_SUFFIX}", room_prefix="C3", elective_room_map=elective_room_map)
    else:
        add_csv_legend_block(ws3, str(courses_v_path), f"CSE {SEM5_SUFFIX}", room_prefix="C3", elective_room_map=elective_room_map)
    combined_v_courses = (c5_block1 or []) + (c5_block2 or [])
    merge_and_color(ws3, combined_v_courses)

    # --- DSAI upper semester ---
    ws8 = wb.create_sheet(f"DSAI-{SEM5_SUFFIX} Timetable")
    d5f_v, d5s_v = split(coursesDSAI_V)
    dsai5_block1 = generate(d5f_v, ws8, f"DSAI-{SEM5_SUFFIX} First Half", seed+18, sync_sem5, room_prefix='C4', elective_room_map=elective_room_map, room_busy_global=global_room_busy,year_tag=YEAR5_TAG,semester_half=1)
    dsai5_block2 = generate(d5s_v, ws8, f"DSAI-{SEM5_SUFFIX} Second Half", seed+19, sync_sem5, room_prefix='C4', elective_room_map=elective_room_map, room_busy_global=global_room_busy,year_tag=YEAR5_TAG,semester_half=2)
    add_csv_legend_block(ws8, str(courses_dsai_v_path), f"DSAI {SEM5_SUFFIX}", room_prefix="C4", elective_room_map=elective_room_map)
    combined_dsai5_courses = (dsai5_block1 or []) + (dsai5_block2 or [])
    merge_and_color(ws8, combined_dsai5_courses)

    # --- ECE upper semester ---
    ws10 = wb.create_sheet(f"ECE-{SEM5_SUFFIX} Timetable")
    e5f_v, e5s_v = split(coursesECE_V)
    ece5_block1 = generate(e5f_v, ws10, f"ECE-{SEM5_SUFFIX} First Half", seed+22, sync_sem5, room_prefix='C4', elective_room_map=elective_room_map, room_busy_global=global_room_busy,year_tag=YEAR5_TAG,semester_half=1)
    ece5_block2 = generate(e5s_v, ws10, f"ECE-{SEM5_SUFFIX} Second Half", seed+23, sync_sem5, room_prefix='C4', elective_room_map=elective_room_map, room_busy_global=global_room_busy,year_tag=YEAR5_TAG,semester_half=2)
    add_csv_legend_block(ws10, str(courses_ece_v_path), f"ECE {SEM5_SUFFIX}", room_prefix="C4", elective_room_map=elective_room_map)
    combined_ece5_courses = (ece5_block1 or []) + (ece5_block2 or [])
    merge_and_color(ws10, combined_ece5_courses)

    # --- Common 7th Sem (optional) ---
    if coursesVII and courses_vii_path is not None:
        ws6 = wb.create_sheet("COMMON 7TH-SEM Timetable")
        s7f, s7s = split(coursesVII)
        s7_block1 = generate(s7f, ws6, "COMMON 7TH-SEM First Half", seed+14, sync_sem7, room_prefix='C3', elective_room_map=elective_room_map, room_busy_global=global_room_busy, year_tag=7, semester_half=1)
        s7_block2 = generate(s7s, ws6, "COMMON 7TH-SEM Second Half", seed+15, sync_sem7, room_prefix='C3', elective_room_map=elective_room_map, room_busy_global=global_room_busy, year_tag=7, semester_half=2)
        add_csv_legend_block(ws6, str(courses_vii_path), "7TH SEM", room_prefix="C3", elective_room_map=elective_room_map)
        combined_7_courses = (s7_block1 or []) + (s7_block2 or [])
        merge_and_color(ws6, combined_7_courses)

    output_path = save_workbook_with_fallback(
        wb, BASE_DIR / "Balanced_Timetable_latest.xlsx"
    )
    print("✅ Evenly balanced timetable saved in", output_path)
