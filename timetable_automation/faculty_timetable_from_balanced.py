import argparse
import difflib
import json
import re
import subprocess
import sys
from collections import defaultdict
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side


DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
DAY_ORDER = {d: i for i, d in enumerate(DAYS)}

thin = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

MANUAL_FACULTY_ALIASES = {
    "anushree": "anusree kini",
    "dibyadyuti": "dibyajyoti guha",
    "shirishkumar layek": "shirshendu l",
    "chimnayananda a": "chinmayananda",
    "jagadisha d n": "jagadish d n",
    "pavan kumar": "pavan kumar c",
    "pramod": "pramod yelmewad",
    "suvadip": "suvadip hazra",
    "utkarsh k": "utkarsh khaire",
    "sunil c k": "sunil ck",
    "sunil kumar p v": "sunil kumar pv",
    "ashwath babu": "dr aswath babu h",
}


def t2m(text):
    h, m = map(int, text.split(":"))
    return h * 60 + m


def load_slot_keys(path):
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    slots = data["time_slots"]
    slots = sorted(slots, key=lambda s: t2m(s["start"]))
    return [f"{s['start']}-{s['end']}" for s in slots]


def slot_start(slot_key):
    return t2m(slot_key.split("-", 1)[0])


def slot_end(slot_key):
    return t2m(slot_key.split("-", 1)[1])


def normalize_space(text):
    return re.sub(r"\s+", " ", str(text).strip())


def normalize_header(text):
    return re.sub(r"[^a-z0-9]", "", str(text).lower())


def normalize_batch_key(text):
    return re.sub(r"[^A-Z0-9]+", "", str(text).upper())


def split_faculty_names(text):
    if text is None:
        return []
    raw = normalize_space(text)
    if not raw:
        return []
    parts = re.split(r"/|;|&|\band\b", raw, flags=re.I)
    return [normalize_space(p) for p in parts if normalize_space(p)]


def normalize_person_name(text):
    t = str(text).lower()
    t = t.replace(".", " ")
    t = re.sub(r"\b(dr|prof|mr|ms|mrs)\b", " ", t)
    t = re.sub(r"[^a-z0-9 ]", " ", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t


def significant_tokens(name):
    toks = [t for t in normalize_person_name(name).split() if len(t) > 1]
    return toks


def build_faculty_directory(faculty_csv):
    df = pd.read_csv(faculty_csv)
    rows = []
    exact = {}
    for _, r in df.iterrows():
        fid = normalize_space(r.get("Faculty_ID", ""))
        name = normalize_space(r.get("Name", ""))
        norm = normalize_person_name(name)
        compact = norm.replace(" ", "")
        toks = significant_tokens(name)
        rec = {
            "id": fid,
            "name": name,
            "norm": norm,
            "compact": compact,
            "tokens": toks,
        }
        rows.append(rec)
        if norm and norm not in exact:
            exact[norm] = rec
    return rows, exact


def resolve_faculty(raw_name, faculty_rows, exact_map):
    raw_name = normalize_space(raw_name)
    if not raw_name:
        return None

    alias_key = normalize_person_name(raw_name)
    if alias_key in MANUAL_FACULTY_ALIASES:
        alias_target = normalize_person_name(MANUAL_FACULTY_ALIASES[alias_key])
        if alias_target in exact_map:
            return exact_map[alias_target]

    norm = normalize_person_name(raw_name)
    if norm in exact_map:
        return exact_map[norm]

    raw_tokens = significant_tokens(raw_name)
    if raw_tokens:
        subset_matches = []
        for rec in faculty_rows:
            rec_tokens = rec["tokens"]
            if all(any(rt == t or t.startswith(rt) or rt.startswith(t) for t in rec_tokens) for rt in raw_tokens):
                subset_matches.append(rec)
        if len(subset_matches) == 1:
            return subset_matches[0]

        first_matches = []
        first = raw_tokens[0]
        for rec in faculty_rows:
            if any(tok == first or tok.startswith(first) or first.startswith(tok) for tok in rec["tokens"]):
                first_matches.append(rec)
        if len(first_matches) == 1:
            return first_matches[0]

    compact = norm.replace(" ", "")
    if compact:
        scored = []
        for rec in faculty_rows:
            score = difflib.SequenceMatcher(None, compact, rec["compact"]).ratio()
            scored.append((score, rec))
        scored.sort(key=lambda x: x[0], reverse=True)
        if scored:
            best_score, best_rec = scored[0]
            second_score = scored[1][0] if len(scored) > 1 else 0.0
            if best_score >= 0.88 and (best_score - second_score) >= 0.03:
                return best_rec
    return None


def get_merged_parent_lookup(ws):
    lookup = {}
    for rng in ws.merged_cells.ranges:
        minr, minc, maxr, maxc = rng.min_row, rng.min_col, rng.max_row, rng.max_col
        for r in range(minr, maxr + 1):
            for c in range(minc, maxc + 1):
                lookup[(r, c)] = (minr, minc)
    return lookup


def cell_value(ws, row, col, merged_lookup):
    top = merged_lookup.get((row, col))
    if top is None:
        return ws.cell(row, col).value
    return ws.cell(top[0], top[1]).value


def parse_legend_sections(ws):
    legends = {}
    r = 1
    while r <= ws.max_row:
        first = ws.cell(r, 1).value
        if isinstance(first, str) and first.strip().lower().startswith("legend -"):
            legend_name = normalize_space(first.split("-", 1)[1])
            hdr_row = r + 1
            header_to_col = {}
            for c in range(1, ws.max_column + 1):
                hv = ws.cell(hdr_row, c).value
                if hv is None:
                    continue
                header_to_col[normalize_header(hv)] = c

            code_col = next((v for k, v in header_to_col.items() if k.startswith("coursecode")), 1)
            title_col = next((v for k, v in header_to_col.items() if k.startswith("coursetitle")), 2)
            fac_col = next((v for k, v in header_to_col.items() if k == "faculty"), 4)
            sem_col = next((v for k, v in header_to_col.items() if k.startswith("semesterhalf")), 5)
            elec_col = next((v for k, v in header_to_col.items() if k == "elective"), 6)
            basket_col = next((v for k, v in header_to_col.items() if "electivebasket" in k), 7)

            rows = []
            rr = hdr_row + 1
            while rr <= ws.max_row:
                first_cell = ws.cell(rr, 1).value
                if isinstance(first_cell, str) and first_cell.strip().lower().startswith("legend -"):
                    break
                if (
                    isinstance(first_cell, str)
                    and re.search(r"(first|second)\s+half$", first_cell.strip(), flags=re.I)
                    and str(ws.cell(rr + 1, 1).value).strip() == "Day"
                ):
                    break
                code_val = normalize_space(ws.cell(rr, code_col).value or "")
                if not code_val:
                    rr += 1
                    continue
                rows.append(
                    {
                        "code": code_val.upper(),
                        "title": normalize_space(ws.cell(rr, title_col).value or ""),
                        "faculty_raw": normalize_space(ws.cell(rr, fac_col).value or ""),
                        "semester_half": normalize_space(ws.cell(rr, sem_col).value or ""),
                        "elective_raw": normalize_space(ws.cell(rr, elec_col).value or ""),
                        "basket_raw": normalize_space(ws.cell(rr, basket_col).value or ""),
                    }
                )
                rr += 1
            legends[legend_name] = rows
            r = rr
            continue
        r += 1
    return legends


def parse_schedule_blocks(ws, slot_keys):
    blocks = []
    merged_lookup = get_merged_parent_lookup(ws)
    for r in range(1, ws.max_row):
        first = ws.cell(r, 1).value
        if not isinstance(first, str):
            continue
        label = normalize_space(first)
        m = re.search(r"(First|Second)\s+Half$", label, flags=re.I)
        if not m:
            continue
        if normalize_space(ws.cell(r + 1, 1).value or "") != "Day":
            continue
        half = m.group(1).lower()
        batch_label = re.sub(r"\s+(First|Second)\s+Half$", "", label, flags=re.I).strip()
        header_row = r + 1
        day_rows = {}
        rr = header_row + 1
        while rr <= ws.max_row:
            dv = normalize_space(ws.cell(rr, 1).value or "")
            if dv not in DAYS:
                break
            day_rows[dv] = rr
            rr += 1
        slot_cols = {}
        for c in range(2, ws.max_column + 1):
            hv = normalize_space(ws.cell(header_row, c).value or "")
            if hv in slot_keys:
                slot_cols[hv] = c
        blocks.append(
            {
                "block_title": label,
                "batch_label": batch_label,
                "half": half,
                "day_rows": day_rows,
                "slot_cols": slot_cols,
                "merged_lookup": merged_lookup,
            }
        )
    return blocks


def legend_for_block(batch_label, legends):
    if not legends:
        return []
    batch_key = normalize_batch_key(batch_label)
    scored = []
    for legend_name, rows in legends.items():
        lk = normalize_batch_key(legend_name)
        if not lk:
            continue
        exact = 1 if lk == batch_key else 0
        overlap = len(set(re.findall(r"[A-Z0-9]+", batch_key)) & set(re.findall(r"[A-Z0-9]+", lk)))
        scored.append((exact, overlap, len(lk), legend_name, rows))
    if not scored:
        return []
    scored.sort(reverse=True)
    return scored[0][4]


def code_from_cell(cell_text):
    text = normalize_space(cell_text)
    if not text:
        return ""
    first = re.split(r"[\s(]", text, maxsplit=1)[0]
    return first.strip().upper()


def canonical_course_code(code):
    code = normalize_space(code).upper()
    if re.match(r"^[A-Z]{2,}\d+[A-Z]?$", code):
        if code.endswith("T") and len(code) > 1 and code[-2].isdigit():
            return code[:-1]
    return code


def parse_elective_bucket(code_text):
    m = re.match(r"^\s*ELECTIVE\s*([0-9]+)", str(code_text).upper())
    if not m:
        return None
    return m.group(1)


def semester_bucket(value):
    v = normalize_space(value).lower()
    if v in {"1", "first", "first half", "h1"}:
        return "first"
    if v in {"2", "second", "second half", "h2"}:
        return "second"
    return "full"


def elective_flag(value):
    v = normalize_space(value).lower()
    return v in {"1", "yes", "true", "y"}


def build_events(input_workbook, slot_keys, faculty_rows, exact_map):
    events_by_faculty = defaultdict(list)
    unmatched_names = set()

    wb = load_workbook(input_workbook, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        legends = parse_legend_sections(ws)
        blocks = parse_schedule_blocks(ws, slot_keys)
        for block in blocks:
            legend_rows = legend_for_block(block["batch_label"], legends)
            for day, row_idx in block["day_rows"].items():
                for slot in slot_keys:
                    col = block["slot_cols"].get(slot)
                    if col is None:
                        continue
                    raw_cell = cell_value(ws, row_idx, col, block["merged_lookup"])
                    text = normalize_space(raw_cell or "")
                    if not text:
                        continue

                    code = code_from_cell(text)
                    if not code:
                        continue

                    bucket = parse_elective_bucket(code)
                    if bucket is not None:
                        selected = []
                        for lr in legend_rows:
                            if not elective_flag(lr["elective_raw"]):
                                continue
                            lr_bucket = normalize_space(lr["basket_raw"])
                            if lr_bucket != bucket:
                                continue
                            sb = semester_bucket(lr["semester_half"])
                            if sb in {"full", block["half"]}:
                                selected.append(lr)
                        if not selected:
                            continue
                        for lr in selected:
                            for raw_fac in split_faculty_names(lr["faculty_raw"]):
                                rec = resolve_faculty(raw_fac, faculty_rows, exact_map)
                                if rec is None:
                                    unmatched_names.add(raw_fac)
                                    continue
                                events_by_faculty[rec["name"]].append(
                                    {
                                        "day": day,
                                        "slot": slot,
                                        "batch": block["block_title"],
                                        "course_code": lr["code"],
                                        "subject": lr["title"] or lr["code"],
                                    }
                                )
                        continue

                    c_code = canonical_course_code(code)
                    matching_rows = [lr for lr in legend_rows if canonical_course_code(lr["code"]) == c_code]
                    if not matching_rows:
                        continue
                    for lr in matching_rows:
                        for raw_fac in split_faculty_names(lr["faculty_raw"]):
                            rec = resolve_faculty(raw_fac, faculty_rows, exact_map)
                            if rec is None:
                                unmatched_names.add(raw_fac)
                                continue
                            events_by_faculty[rec["name"]].append(
                                {
                                    "day": day,
                                    "slot": slot,
                                    "batch": block["block_title"],
                                    "course_code": lr["code"],
                                    "subject": lr["title"] or lr["code"],
                                }
                            )
    return events_by_faculty, unmatched_names


def merge_contiguous_slots(events):
    grouped = defaultdict(list)
    for e in events:
        key = (e["day"], e["batch"], e["course_code"], e["subject"])
        grouped[key].append(e["slot"])

    merged = []
    for key, slots in grouped.items():
        slots = sorted(set(slots), key=slot_start)
        day, batch, course_code, subject = key
        if not slots:
            continue
        cur_start = slots[0].split("-", 1)[0]
        cur_end = slots[0].split("-", 1)[1]
        for s in slots[1:]:
            s_start, s_end = s.split("-", 1)
            if t2m(s_start) == t2m(cur_end):
                cur_end = s_end
            else:
                merged.append(
                    {
                        "day": day,
                        "time": f"{cur_start}-{cur_end}",
                        "batch": batch,
                        "course_code": course_code,
                        "subject": subject,
                    }
                )
                cur_start, cur_end = s_start, s_end
        merged.append(
            {
                "day": day,
                "time": f"{cur_start}-{cur_end}",
                "batch": batch,
                "course_code": course_code,
                "subject": subject,
            }
        )
    merged.sort(key=lambda x: (DAY_ORDER.get(x["day"], 99), t2m(x["time"].split("-", 1)[0]), x["batch"], x["course_code"]))
    return merged


def safe_sheet_title(text):
    t = re.sub(r'[:\\/*?\[\]]', "_", text)
    t = t.strip()
    return t[:31] if t else "Sheet"


def apply_basic_style(ws):
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if cell.row == 1:
                cell.font = Font(bold=True)
    for col in ws.columns:
        letter = col[0].column_letter
        max_len = 0
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[letter].width = min(max_len + 2, 55)


def build_output(
    faculty_rows,
    merged_events_by_faculty,
    unmatched_names,
    output_workbook,
):
    wb = Workbook()
    summary = wb.active
    summary.title = "Faculty_Summary"
    summary.append(["Faculty_ID", "Faculty_Name", "Day", "Time", "Batch", "Course_Code", "Subject"])

    no_classes = []
    used_titles = set()

    for rec in faculty_rows:
        name = rec["name"]
        fid = rec["id"]
        events = merged_events_by_faculty.get(name, [])
        if not events:
            no_classes.append([fid, name])
        for e in events:
            summary.append([fid, name, e["day"], e["time"], e["batch"], e["course_code"], e["subject"]])

        base = safe_sheet_title(f"{fid}_{name}")
        title = base
        idx = 1
        while title in used_titles:
            suffix = f"_{idx}"
            title = safe_sheet_title(base[: 31 - len(suffix)] + suffix)
            idx += 1
        used_titles.add(title)

        ws = wb.create_sheet(title)
        ws.append(["Day", "Time", "Batch", "Course_Code", "Subject"])
        if events:
            for e in events:
                ws.append([e["day"], e["time"], e["batch"], e["course_code"], e["subject"]])
        else:
            ws.append(["-", "-", "-", "-", "No classes assigned in balanced timetable"])
        apply_basic_style(ws)

    ws_no = wb.create_sheet("No_Classes")
    ws_no.append(["Faculty_ID", "Faculty_Name"])
    for row in no_classes:
        ws_no.append(row)
    apply_basic_style(ws_no)

    ws_un = wb.create_sheet("Unmapped_Names")
    ws_un.append(["Raw_Faculty_Name_From_Legend"])
    for name in sorted(unmatched_names):
        ws_un.append([name])
    apply_basic_style(ws_un)

    apply_basic_style(summary)
    wb.save(output_workbook)


def run_best_of_n_generation(
    trials,
    seed_start,
    output_workbook,
    report_json,
    slots_json,
    keep_temp,
    temp_dir,
    room_check_script,
    skip_room_check,
):
    root = Path(__file__).resolve().parents[1]
    optimizer = root / "timetable_automation" / "optimize_balanced_timetable.py"
    cmd = [
        sys.executable,
        str(optimizer),
        "--trials",
        str(trials),
        "--output",
        str(output_workbook),
        "--report-json",
        str(report_json),
        "--slots-json",
        str(slots_json),
        "--temp-dir",
        str(temp_dir),
        "--room-check-script",
        str(room_check_script),
    ]
    if seed_start is not None:
        cmd.extend(["--seed-start", str(seed_start)])
    if keep_temp:
        cmd.append("--keep-temp")
    if skip_room_check:
        cmd.append("--skip-room-check")

    proc = subprocess.run(cmd, cwd=root, capture_output=True, text=True)
    if proc.stdout.strip():
        print(proc.stdout.strip())
    if proc.returncode != 0:
        raise RuntimeError(
            "Best-of-N generation failed.\n"
            f"Command: {' '.join(cmd)}\n"
            f"STDOUT:\n{proc.stdout}\n"
            f"STDERR:\n{proc.stderr}"
        )


def main():
    parser = argparse.ArgumentParser(
        description="Generate faculty-wise timetable from Balanced_Timetable_latest.xlsx"
    )
    parser.add_argument(
        "--input",
        default="Balanced_Timetable_latest.xlsx",
        help="Input balanced timetable workbook",
    )
    parser.add_argument(
        "--faculty-csv",
        default="data/Faculty.csv",
        help="Faculty master list CSV",
    )
    parser.add_argument(
        "--slots-json",
        default="data/time_slots.json",
        help="Time slots JSON path",
    )
    parser.add_argument(
        "--output",
        default="Faculty_Timetable_from_Balanced.xlsx",
        help="Output faculty timetable workbook",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse and report counts without writing workbook",
    )
    parser.add_argument(
        "--best-of",
        type=int,
        default=0,
        help="If >0, run best-of-N timetable generation+validation before faculty extraction",
    )
    parser.add_argument(
        "--best-seed-start",
        type=int,
        default=None,
        help="Optional starting seed for best-of-N generation",
    )
    parser.add_argument(
        "--best-report-json",
        default="balanced_timetable_validation.json",
        help="Validation report JSON path for best-of-N generation",
    )
    parser.add_argument(
        "--best-keep-temp",
        action="store_true",
        help="Keep temporary seed trial workbooks/reports from best-of-N generation",
    )
    parser.add_argument(
        "--best-temp-dir",
        default="timetable_automation/.opt_tmp",
        help="Temporary directory used during best-of-N seed trials",
    )
    parser.add_argument(
        "--best-room-check-script",
        default="check_room_clashes.py",
        help="Room clash checker script path used during best-of-N scoring",
    )
    parser.add_argument(
        "--best-skip-room-check",
        action="store_true",
        help="Skip room clash checker in best-of-N scoring",
    )
    args = parser.parse_args()

    if args.best_of > 0:
        run_best_of_n_generation(
            trials=args.best_of,
            seed_start=args.best_seed_start,
            output_workbook=args.input,
            report_json=args.best_report_json,
            slots_json=args.slots_json,
            keep_temp=args.best_keep_temp,
            temp_dir=args.best_temp_dir,
            room_check_script=args.best_room_check_script,
            skip_room_check=args.best_skip_room_check,
        )

    slot_keys = load_slot_keys(args.slots_json)
    faculty_rows, exact_map = build_faculty_directory(args.faculty_csv)
    events_by_faculty, unmatched_names = build_events(args.input, slot_keys, faculty_rows, exact_map)

    merged_events_by_faculty = {}
    for rec in faculty_rows:
        fname = rec["name"]
        merged_events_by_faculty[fname] = merge_contiguous_slots(events_by_faculty.get(fname, []))

    if args.dry_run:
        assigned = sum(1 for rec in faculty_rows if merged_events_by_faculty.get(rec["name"]))
        total_events = sum(len(v) for v in merged_events_by_faculty.values())
        print(f"Faculty in CSV: {len(faculty_rows)}")
        print(f"Faculty with at least one class: {assigned}")
        print(f"Total merged faculty events: {total_events}")
        print(f"Unmapped faculty names from legend: {len(unmatched_names)}")
        if unmatched_names:
            print("Unmapped names:")
            for name in sorted(unmatched_names):
                print(f" - {name}")
        return

    build_output(
        faculty_rows=faculty_rows,
        merged_events_by_faculty=merged_events_by_faculty,
        unmatched_names=unmatched_names,
        output_workbook=args.output,
    )
    print(f"{args.output} generated successfully.")


if __name__ == "__main__":
    main()
