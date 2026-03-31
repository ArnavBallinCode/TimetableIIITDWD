import argparse
import datetime as dt
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

FILE_PATH = "CourseCode&Name.csv"
OUTPUT_FILE = "Exam_Timetable.xlsx"


def parse_start_date(value: str) -> dt.date:
    return dt.datetime.strptime(value, "%Y-%m-%d").date()


def get_shift(batch_name: str):
    b = batch_name.upper()
    if "1ST" in b or "3RD" in b:
        return "Morning (10:00 AM – 11:30 AM)"
    return "Evening (03:00 PM – 04:30 PM)"


def parse_courses(file_path: Path):
    courses = {}
    current_batch = None
    current_courses = []

    with file_path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()

            if line.startswith("BATCH"):
                if current_batch and current_courses:
                    courses[current_batch] = current_courses
                current_batch = line.replace("BATCH", "").replace(":", "").strip()
                current_courses = []
            elif line and "," in line:
                code, name = [x.strip() for x in line.split(",", 1)]
                current_courses.append((code, name))
            elif not line:
                if current_batch and current_courses:
                    courses[current_batch] = current_courses
                    current_batch, current_courses = None, []

    if current_batch and current_courses:
        courses[current_batch] = current_courses

    return courses


def generate_exam_dates(num_days, start_date):
    dates = []
    d = start_date
    while len(dates) < num_days:
        if d.weekday() < 5:
            dates.append(d)
        d += dt.timedelta(days=1)
    return dates


def generate_timetable_dataframe(courses, start_date: dt.date):
    max_exams = max(len(c) for c in courses.values())
    dates = generate_exam_dates(max_exams + 5, start_date)
    records = []

    for batch_name, clist in courses.items():
        shift = get_shift(batch_name)
        normal_courses = [(code, name) for code, name in clist if "ELECTIVE" not in code.upper()]
        elective_courses = [(code, name) for code, name in clist if "ELECTIVE" in code.upper()]
        date_idx = 0
        used_dates = set()

        for code, name in normal_courses:
            exam_date = dates[date_idx]
            used_dates.add(exam_date)
            records.append(
                {
                    "Batch": batch_name,
                    "Date": exam_date,
                    "Day": exam_date.strftime("%A"),
                    "Shift": shift,
                    "CourseCode": code,
                    "CourseName": name,
                }
            )
            date_idx += 1

        if elective_courses:
            while date_idx < len(dates) and dates[date_idx] in used_dates:
                date_idx += 1
            elective_date = dates[date_idx]
            for code, name in elective_courses:
                records.append(
                    {
                        "Batch": batch_name,
                        "Date": elective_date,
                        "Day": elective_date.strftime("%A"),
                        "Shift": shift,
                        "CourseCode": code,
                        "CourseName": name,
                    }
                )

    df = pd.DataFrame(records).sort_values(by=["Batch", "Date"]).reset_index(drop=True)
    df["Date"] = pd.to_datetime(df["Date"])
    df["Date_str"] = df["Date"].dt.strftime("%d-%b-%Y")
    return df


def write_timetable_excel(df, output_file: Path):
    wb = Workbook()
    ws_master = wb.active
    ws_master.title = "Master_Timetable"
    headers = ["Batch", "Date", "Day", "Shift", "Course Code", "Course Name"]
    for col, head in enumerate(headers, start=1):
        ws_master.cell(row=1, column=col, value=head)
        ws_master.cell(row=1, column=col).font = Font(bold=True)
        ws_master.cell(row=1, column=col).alignment = Alignment(horizontal="center")

    for i, row in df.iterrows():
        row_data = [row["Batch"], row["Date_str"], row["Day"], row["Shift"], row["CourseCode"], row["CourseName"]]
        for j, val in enumerate(row_data, start=1):
            ws_master.cell(row=i + 2, column=j, value=val)
            ws_master.cell(row=i + 2, column=j).alignment = Alignment(horizontal="center")

    for batch in df["Batch"].unique():
        sub_df = df[df["Batch"] == batch].sort_values(by="Date")
        ws = wb.create_sheet(title=batch[:30])
        shift_for_sheet = get_shift(batch)

        ws.merge_cells("A1:F1")
        ws["A1"] = f"Exam Timetable - {batch} ({shift_for_sheet})"
        ws["A1"].alignment = Alignment(horizontal="center")
        ws["A1"].font = Font(bold=True, size=13)

        for idx, col in enumerate(["Date", "Day", "Course Code", "Course Name"], start=1):
            ws.cell(row=2, column=idx, value=col).font = Font(bold=True)
            ws.cell(row=2, column=idx).alignment = Alignment(horizontal="center")

        for i, r in enumerate(sub_df.itertuples(), start=3):
            ws.cell(row=i, column=1, value=r.Date_str)
            ws.cell(row=i, column=2, value=r.Day)
            ws.cell(row=i, column=3, value=r.CourseCode)
            ws.cell(row=i, column=4, value=r.CourseName)
            for c in range(1, 5):
                ws.cell(row=i, column=c).alignment = Alignment(horizontal="center")

        thin = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        for row in ws.iter_rows(min_row=2, max_row=i, min_col=1, max_col=4):
            for cell in row:
                cell.border = thin

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

    wb.save(output_file)


def build_parser():
    parser = argparse.ArgumentParser(description="Generate exam timetable from CourseCode&Name.csv format")
    parser.add_argument("--input-file", default=FILE_PATH, help=f"Input file path (default: {FILE_PATH})")
    parser.add_argument("--output-file", default=OUTPUT_FILE, help=f"Output xlsx path (default: {OUTPUT_FILE})")
    parser.add_argument(
        "--start-date",
        type=parse_start_date,
        help="Exam start date in YYYY-MM-DD format (default: today's date)",
    )
    return parser


def main(argv=None):
    args = build_parser().parse_args(argv)
    start_date = args.start_date or dt.date.today()
    if start_date < dt.date.today():
        print(f"⚠ Warning: start date {start_date.isoformat()} is in the past.")

    input_file = Path(args.input_file)
    if not input_file.exists():
        raise FileNotFoundError(f"Input file not found: {input_file}")

    courses = parse_courses(input_file)
    if not courses:
        raise ValueError("No batches/courses found in input file.")

    print(f"✅ Parsed {len(courses)} batches successfully!")
    df = generate_timetable_dataframe(courses, start_date)
    write_timetable_excel(df, Path(args.output_file))
    print(f"🎯 Timetable created successfully → {args.output_file}")


if __name__ == "__main__":
    main()
